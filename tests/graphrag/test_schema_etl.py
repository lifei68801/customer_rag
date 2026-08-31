from __future__ import annotations

from pathlib import Path

import aiosqlite
import pytest

from app.graphrag.etl_projection import DuplicateNodeKeyError
from app.graphrag.etl_stable_code_registry import (
    allocate_stable_code,
    ensure_stable_code_registry_schema,
    lookup_stable_code,
)
from app.graphrag.ontology_categories import ExtraFieldSpec, create_term_type
from app.graphrag.ontology_constraints import add_allowed_combination
from app.graphrag.ontology_lifecycle import checkout_draft, confirm_ontology, ensure_ontology_schema
from app.graphrag.ontology_relations import create_relation_type
from app.graphrag.schema_etl import (
    DuplicateEntityMappingError,
    SchemaETLNotConfirmedError,
    SweepSafetyValveError,
    run_schema_etl,
)
from app.graphrag.schema_etl_config import (
    AllocatedCodeNodeKeyPart,
    ColumnNodeKeyPart,
    EntityMapping,
    RelationMapping,
    SchemaETLConfig,
)
from app.graphrag.terms_store import ensure_terms_schema, get_term, list_terms, upsert_term_with_node_key

pytestmark = pytest.mark.anyio


class FakeGraphClient:
    def __init__(self) -> None:
        self.synced: list[str] = []
        self.merged: list[tuple[str, str, str]] = []
        self.deleted_nodes: list[str] = []
        self.stale_sweeps: list[tuple[str, str]] = []

    async def sync_term(self, term) -> None:
        self.synced.append(term.node_key)

    async def merge_relation(
        self, *, subject_standard_name, object_standard_name, relation_type,
        source, tenant_id, provenance, recorded_at,
    ) -> None:
        self.merged.append((subject_standard_name, object_standard_name, relation_type))

    async def delete_term_node(self, *, tenant_id: str, node_key: str) -> None:
        self.deleted_nodes.append(node_key)

    async def delete_stale_relations_by_source(
        self, source: str, *, tenant_id: str, before_recorded_at: str
    ) -> int:
        self.stale_sweeps.append((source, before_recorded_at))
        # 假客户端不真的维护边集合，返回 0；真实计数由 Neo4j 侧的实现负责。
        return 0


async def _confirmed_conn() -> aiosqlite.Connection:
    conn = await aiosqlite.connect(":memory:")
    await ensure_terms_schema(conn)
    await ensure_ontology_schema(conn)
    await ensure_stable_code_registry_schema(conn)
    await create_term_type(
        conn, tenant_id="muji", value="Product",
        extra_fields=[ExtraFieldSpec(name="md_no", value_type="string")],
    )
    await create_term_type(conn, tenant_id="muji", value="SKU")
    await create_term_type(conn, tenant_id="muji", value="VariantValue")
    await checkout_draft(conn, "muji")
    await create_relation_type(
        conn, "muji", relation_type="HAS_SKU", example_phrase="Product HAS_SKU SKU",
    )
    await create_relation_type(
        conn, "muji", relation_type="HAS_VARIANT_VALUE",
        example_phrase="VariantValue HAS_VARIANT_VALUE VariantValue",
    )
    await add_allowed_combination(
        conn, "muji", subject_term_type="Product", relation_type="HAS_SKU", object_term_type="SKU",
    )
    await add_allowed_combination(
        conn, "muji", subject_term_type="VariantValue", relation_type="HAS_VARIANT_VALUE",
        object_term_type="VariantValue",
    )
    await confirm_ontology(conn, "muji")
    return conn


async def test_run_schema_etl_raises_when_schema_not_confirmed():
    conn = await aiosqlite.connect(":memory:")
    await ensure_terms_schema(conn)
    await ensure_ontology_schema(conn)
    await ensure_stable_code_registry_schema(conn)
    config = SchemaETLConfig(tenant_id="muji", entities=[], relations=[])

    with pytest.raises(SchemaETLNotConfirmedError):
        await run_schema_etl(conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=Path("."))


async def test_run_schema_etl_writes_entities_and_relations(tmp_path):
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name,md_no\n1001,圆角收纳盒,A123\n", encoding="utf-8"
    )
    (tmp_path / "skus.csv").write_text(
        "jan,product_group_id\n4901234567890,1001\n", encoding="utf-8"
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
            EntityMapping(
                term_type="SKU", source_file="skus.csv",
                standard_name_parts=["jan"],
                node_key_parts=[ColumnNodeKeyPart(column="jan")],
                field_mappings={},
            ),
        ],
        relations=[
            RelationMapping(
                relation_type="HAS_SKU", source_file="skus.csv",
                subject_term_type="Product", object_term_type="SKU",
            ),
        ],
    )
    graph_client = FakeGraphClient()

    report = await run_schema_etl(conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path)

    assert report.entities_written == 2
    assert report.entities_skipped == 0
    assert report.relations_written == 1
    product = await get_term(conn, tenant_id="muji", standard_name="圆角收纳盒")
    assert product.node_key == "Product:1001"
    assert product.extra_properties == {"md_no": "A123"}
    assert "Product:1001" in graph_client.synced
    assert "SKU:4901234567890" in graph_client.synced
    assert ("Product:1001", "SKU:4901234567890", "HAS_SKU") in graph_client.merged


async def test_run_schema_etl_skips_bad_row_and_reports_it(tmp_path):
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name,md_no\n"
        "1001,圆角收纳盒,A123\n"
        ",没有ID的商品,B456\n",  # 第二行缺 product_group_id
        encoding="utf-8",
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
        ],
        relations=[],
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.entities_written == 1
    assert report.entities_skipped == 1
    assert len(report.skipped_rows) == 1
    assert "products.csv" in report.skipped_rows[0].source_file


async def test_run_schema_etl_rerun_is_idempotent(tmp_path):
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name,md_no\n1001,圆角收纳盒,A123\n", encoding="utf-8"
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
        ],
        relations=[],
    )

    await run_schema_etl(conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path)
    await run_schema_etl(conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path)

    from app.graphrag.terms_store import list_terms
    all_terms = await list_terms(conn, tenant_id="muji")
    assert len(all_terms) == 1


async def test_run_schema_etl_skips_bad_row_in_the_middle_and_still_processes_the_rest(tmp_path):
    """坏行出现在文件中间时，它后面的行也必须继续处理——不能因为一行脏数据
    就提前结束整个文件的遍历（设计文档第 6.4 节）。"""
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name,md_no\n"
        "1001,圆角收纳盒,A123\n"
        ",没有ID的商品,B456\n"  # 中间这一行缺 product_group_id
        "1003,亚麻抱枕套,C789\n",
        encoding="utf-8",
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
        ],
        relations=[],
    )
    graph_client = FakeGraphClient()

    report = await run_schema_etl(
        conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
    )

    assert report.entities_written == 2
    assert report.entities_skipped == 1
    assert graph_client.synced == ["Product:1001", "Product:1003"]
    assert (await get_term(conn, tenant_id="muji", standard_name="圆角收纳盒")) is not None
    assert (await get_term(conn, tenant_id="muji", standard_name="亚麻抱枕套")) is not None
    assert report.skipped_rows[0].row_number == 3


async def test_run_schema_etl_unconfirmed_relation_type_skips_only_that_mapping(tmp_path):
    """relation_type 不在已确认 schema 里，只跳过这一个 mapping、记进
    skipped_mappings，其余实体照常写入，不抛异常中断整次运行。"""
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name,md_no\n1001,圆角收纳盒,A123\n", encoding="utf-8"
    )
    (tmp_path / "skus.csv").write_text(
        "jan,product_group_id\n4901234567890,1001\n", encoding="utf-8"
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
            EntityMapping(
                term_type="SKU", source_file="skus.csv",
                standard_name_parts=["jan"],
                node_key_parts=[ColumnNodeKeyPart(column="jan")],
                field_mappings={},
            ),
        ],
        relations=[
            RelationMapping(
                relation_type="NOT_REGISTERED", source_file="skus.csv",
                subject_term_type="Product", object_term_type="SKU",
            ),
        ],
    )
    graph_client = FakeGraphClient()

    report = await run_schema_etl(
        conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
    )

    assert report.entities_written == 2
    assert report.relations_written == 0
    assert graph_client.merged == []
    assert len(report.skipped_mappings) == 1
    assert report.skipped_mappings[0].label == "NOT_REGISTERED"


async def test_run_schema_etl_relation_type_confirmed_but_combination_not_allowed_skips_only_that_mapping(
    tmp_path,
):
    """relation_type 本身已确认，但 (subject_term_type, relation_type,
    object_term_type) 这个组合不在已确认的允许列表里——只跳过这一个
    mapping、记进 skipped_mappings，其余实体照常写入，不抛异常中断整次
    运行。HAS_SKU 在 _confirmed_conn() 里只允许 (Product, HAS_SKU, SKU)
    这一个方向，这里故意把主体/客体类型颠倒过来触发校验。"""
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name,md_no\n1001,圆角收纳盒,A123\n", encoding="utf-8"
    )
    (tmp_path / "skus.csv").write_text(
        "jan,product_group_id\n4901234567890,1001\n", encoding="utf-8"
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
            EntityMapping(
                term_type="SKU", source_file="skus.csv",
                standard_name_parts=["jan"],
                node_key_parts=[ColumnNodeKeyPart(column="jan")],
                field_mappings={},
            ),
        ],
        relations=[
            RelationMapping(
                relation_type="HAS_SKU", source_file="skus.csv",
                subject_term_type="SKU", object_term_type="Product",
            ),
        ],
    )
    graph_client = FakeGraphClient()

    report = await run_schema_etl(
        conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
    )

    assert report.entities_written == 2
    assert report.relations_written == 0
    assert graph_client.merged == []
    assert len(report.skipped_mappings) == 1
    assert report.skipped_mappings[0].label == "HAS_SKU"
    assert "允许列表" in report.skipped_mappings[0].reason


async def test_run_schema_etl_unregistered_term_type_skips_only_that_mapping(tmp_path):
    """term_type 没在该租户注册过时同理：跳过这一个实体 mapping，
    其它实体和关系照常处理。"""
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name,md_no\n1001,圆角收纳盒,A123\n", encoding="utf-8"
    )
    (tmp_path / "skus.csv").write_text(
        "jan,product_group_id\n4901234567890,1001\n", encoding="utf-8"
    )
    (tmp_path / "unknown.csv").write_text("code,name\nX1,未知类型\n", encoding="utf-8")
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="NotRegistered", source_file="unknown.csv",
                standard_name_parts=["name"],
                node_key_parts=[ColumnNodeKeyPart(column="code")],
                field_mappings={},
            ),
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
            EntityMapping(
                term_type="SKU", source_file="skus.csv",
                standard_name_parts=["jan"],
                node_key_parts=[ColumnNodeKeyPart(column="jan")],
                field_mappings={},
            ),
        ],
        relations=[
            RelationMapping(
                relation_type="HAS_SKU", source_file="skus.csv",
                subject_term_type="Product", object_term_type="SKU",
            ),
        ],
    )
    graph_client = FakeGraphClient()

    report = await run_schema_etl(
        conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
    )

    assert report.entities_written == 2
    assert report.relations_written == 1
    assert len(report.skipped_mappings) == 1
    assert report.skipped_mappings[0].label == "NotRegistered"
    assert report.skipped_mappings[0].source_file == "unknown.csv"


async def test_run_schema_etl_unregistered_term_type_with_allocated_code_key_allocates_no_orphan_code(tmp_path):
    """预检要先看一眼 term_type 在不在已确认 schema 里，跳过不在里面的
    mapping、不去扫它的键：scan_entity_node_keys 内部的
    compute_node_key(allow_allocation=True) 会为 AllocatedCodeNodeKeyPart
    真实分配并持久化稳定码，它不检查 term_type 合不合法。重构前 term_type
    校验在 _write_entity_mapping 的行循环之前，打错字的 mapping 从不触发
    compute_node_key；如果预检不加这层判断，即便整次运行完全成功，一个
    term_type 打错字的 mapping 也会往 etl_stable_code_registry 里写入
    永远不会被业务数据引用的孤儿码——这是相对重构前的真实行为倒退。
    term_type 非法本身仍然只由 _write_entity_mapping 判定、记进
    skipped_mappings，预检不抢先 raise，语义不变。"""
    conn = await _confirmed_conn()
    (tmp_path / "unknown.csv").write_text("code,name\nX1,未知类型\n", encoding="utf-8")
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="NotRegistered", source_file="unknown.csv",
                standard_name_parts=["name"],
                node_key_parts=[
                    AllocatedCodeNodeKeyPart(scope_columns=[], raw_value_column="code")
                ],
                field_mappings={},
            ),
        ],
        relations=[],
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert len(report.skipped_mappings) == 1
    assert report.skipped_mappings[0].label == "NotRegistered"
    assert await lookup_stable_code(
        conn, tenant_id="muji", scope="NotRegistered", raw_value="X1"
    ) is None


async def test_run_schema_etl_relation_endpoint_never_written_is_skipped_not_ghost_merged(tmp_path):
    """关系文件引用了一个实体文件里从来没出现过的原始值时，必须跳过这一行——
    绝不能顺手给它分配一个新的稳定码、MERGE 出一个没有对应 Term 记录的幽灵节点。"""
    conn = await _confirmed_conn()
    (tmp_path / "variant_values.csv").write_text(
        "dim_code,raw_value\ndim_007,抹茶\n", encoding="utf-8"
    )
    (tmp_path / "variant_links.csv").write_text(
        "dim_code,raw_value\ndim_007,从没出现过的值\n", encoding="utf-8"
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="VariantValue", source_file="variant_values.csv",
                standard_name_parts=["raw_value"],
                node_key_parts=[
                    AllocatedCodeNodeKeyPart(scope_columns=["dim_code"], raw_value_column="raw_value")
                ],
                field_mappings={},
            ),
        ],
        relations=[
            RelationMapping(
                relation_type="HAS_VARIANT_VALUE", source_file="variant_links.csv",
                subject_term_type="VariantValue", object_term_type="VariantValue",
            ),
        ],
    )
    graph_client = FakeGraphClient()

    report = await run_schema_etl(
        conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
    )

    assert report.entities_written == 1
    assert report.relations_written == 0
    assert report.relations_skipped == 1
    assert graph_client.merged == []
    # 关系路径没有消耗计数位：scope 下只有实体路径分配过的那一个码
    next_code = await allocate_stable_code(
        conn, tenant_id="muji", scope="VariantValue:dim_007", raw_value="从没出现过的值"
    )
    assert next_code == "00002"


async def test_run_schema_etl_column_key_endpoint_never_written_is_skipped_not_ghost_merged(
    tmp_path,
):
    """普通列节点键也要有和稳定码同样的守卫：关系文件引用了一个实体文件里
    从来没写成功过的值时，必须跳过这一行，不能 MERGE 出一个只有 node_key、
    没有 type/standard_name 的幽灵节点。

    这是真实发生过的事故：demo 租户的 `类目:Coffee` 和 `销量:1000` 两个实体
    因为当时的唯一索引冲突写入失败，关系写入这一趟却照写不误，在图里留下两个
    裸 :Term 节点挂着 16 条边。稳定码路径早就有守卫（见上一个测试），
    ColumnNodeKeyPart 路径没有。
    """
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_text("name,md_no\nP1,MD1\n", encoding="utf-8")
    (tmp_path / "skus.csv").write_text("sku_code\nS1\n", encoding="utf-8")
    (tmp_path / "links.csv").write_text("name,sku_code\nP1,S_NEVER_WRITTEN\n", encoding="utf-8")
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv", standard_name_parts=["name"],
                node_key_parts=[ColumnNodeKeyPart(column="name")], field_mappings={},
            ),
            EntityMapping(
                term_type="SKU", source_file="skus.csv", standard_name_parts=["sku_code"],
                node_key_parts=[ColumnNodeKeyPart(column="sku_code")], field_mappings={},
            ),
        ],
        relations=[
            RelationMapping(
                relation_type="HAS_SKU", source_file="links.csv",
                subject_term_type="Product", object_term_type="SKU",
            ),
        ],
    )
    graph_client = FakeGraphClient()

    report = await run_schema_etl(
        conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
    )

    assert report.entities_written == 2
    assert report.relations_written == 0
    assert report.relations_skipped == 1
    assert graph_client.merged == []


async def test_run_schema_etl_reports_per_type_counts(tmp_path):
    """设计文档第 6.4 节要求汇总报告能按 term_type/relation_type 分别给出
    写入行数，不只是整次运行的总量。"""
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name,md_no\n1001,圆角收纳盒,A123\n", encoding="utf-8"
    )
    (tmp_path / "skus.csv").write_text(
        "jan,product_group_id\n4901234567890,1001\n", encoding="utf-8"
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
            EntityMapping(
                term_type="SKU", source_file="skus.csv",
                standard_name_parts=["jan"],
                node_key_parts=[ColumnNodeKeyPart(column="jan")],
                field_mappings={},
            ),
        ],
        relations=[
            RelationMapping(
                relation_type="HAS_SKU", source_file="skus.csv",
                subject_term_type="Product", object_term_type="SKU",
            ),
        ],
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.written_by_type == {"Product": 1, "SKU": 1, "HAS_SKU": 1}
    assert report.skipped_by_type == {}


async def test_run_schema_etl_reads_gbk_encoded_csv(tmp_path):
    """国内 Excel 导出 CSV 常见默认编码是 GBK，不是 UTF-8——见
    docs/superpowers/specs/2026-08-21-schema-etl-multi-format-upload.md
    决策 6。这里直接写 GBK 编码的字节，不依赖任何自动转码工具，验证读取器
    自己能探测出编码并正确解码出中文列名/值。"""
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_bytes(
        "product_group_id,product_group_name,md_no\n1001,圆角收纳盒,A123\n".encode("gbk")
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
        ],
        relations=[],
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.entities_written == 1
    assert report.entities_skipped == 0
    term = await get_term(conn, tenant_id="muji", standard_name="圆角收纳盒")
    assert term is not None
    assert term.node_key == "Product:1001"


async def test_run_schema_etl_reads_tsv_source_file(tmp_path):
    """TSV 只是分隔符从逗号换成制表符，验证扩展名 .tsv 能被正确识别并用
    制表符分隔解析——见决策 1。"""
    conn = await _confirmed_conn()
    (tmp_path / "products.tsv").write_text(
        "product_group_id\tproduct_group_name\tmd_no\n1001\t圆角收纳盒\tA123\n", encoding="utf-8"
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.tsv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
        ],
        relations=[],
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.entities_written == 1
    term = await get_term(conn, tenant_id="muji", standard_name="圆角收纳盒")
    assert term is not None
    assert term.node_key == "Product:1001"


def _write_xlsx_fixture(path, *, header: list[str], rows: list[list[object]]) -> None:
    """用 openpyxl 生成一个最小的 xlsx 测试夹具文件——openpyxl 既能读也能
    写，不需要额外引入别的库。"""
    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(header)
    for row in rows:
        worksheet.append(row)
    workbook.save(str(path))


def _write_xls_fixture(path, *, header: list[str], rows: list[list[object]]) -> None:
    """用 xlwt 生成一个最小的 xls（旧版二进制 Excel）测试夹具文件——xlrd
    只能读 xls 不能写，xlwt 只能写 xls 不能写 xlsx，两个库分工明确，这里
    只用来造测试数据，生产代码不导入 xlwt。"""
    import xlwt

    workbook = xlwt.Workbook()
    worksheet = workbook.add_sheet("Sheet1")
    for col, name in enumerate(header):
        worksheet.write(0, col, name)
    for row_idx, row in enumerate(rows, start=1):
        for col, value in enumerate(row):
            worksheet.write(row_idx, col, value)
    workbook.save(str(path))


async def test_run_schema_etl_reads_xlsx_source_file(tmp_path):
    conn = await _confirmed_conn()
    _write_xlsx_fixture(
        tmp_path / "products.xlsx",
        header=["product_group_id", "product_group_name", "md_no"],
        rows=[[1001, "圆角收纳盒", "A123"]],
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.xlsx",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
        ],
        relations=[],
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.entities_written == 1
    assert report.entities_skipped == 0
    term = await get_term(conn, tenant_id="muji", standard_name="圆角收纳盒")
    assert term is not None
    assert term.node_key == "Product:1001"
    # xlsx 单元格里 1001 是原生 int，node_key 必须是 "Product:1001" 而不是
    # "Product:1001.0"——验证 convert_excel_cell_to_string 真的被用在了
    # 读取路径上，不是只在 Task 2 的单元测试里孤立存在。
    # 注意：get_term(conn, tenant_id, standard_name) 的真实签名只接受
    # standard_name，不接受 node_key 参数——按 standard_name 查询，
    # 再断言返回对象的 .node_key 字段。


async def test_run_schema_etl_reads_xls_source_file(tmp_path):
    conn = await _confirmed_conn()
    _write_xls_fixture(
        tmp_path / "products.xls",
        header=["product_group_id", "product_group_name", "md_no"],
        rows=[[1001, "圆角收纳盒", "A123"]],
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.xls",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
        ],
        relations=[],
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.entities_written == 1
    term = await get_term(conn, tenant_id="muji", standard_name="圆角收纳盒")
    assert term is not None
    assert term.node_key == "Product:1001"


async def test_run_schema_etl_xlsx_empty_sheet_writes_nothing(tmp_path):
    """只有表头没有数据行的 xlsx，不应该报错，也不应该写入任何实体——跟
    CSV 场景下"只有表头"的行为一致。"""
    conn = await _confirmed_conn()
    _write_xlsx_fixture(
        tmp_path / "products.xlsx",
        header=["product_group_id", "product_group_name", "md_no"],
        rows=[],
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.xlsx",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
        ],
        relations=[],
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.entities_written == 0
    assert report.entities_skipped == 0


async def test_run_schema_etl_xlsx_phantom_trailing_row_is_skipped_not_counted(tmp_path):
    """openpyxl 的 read_only 迭代会按工作表已用范围补齐行数，哪怕某一行
    早就被清空也会产出全空的幽灵行（常见于手工编辑过的 Excel 导出文件）。
    这条用例专门构造一个"先写值、再清空"的行，验证它被安静跳过、不计入
    entities_skipped，而不是被当成脏数据报出来。"""
    conn = await _confirmed_conn()

    from openpyxl import Workbook

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["product_group_id", "product_group_name", "md_no"])
    worksheet.append([1001, "圆角收纳盒", "A123"])
    # 故意先写一个值再清空，让 openpyxl 的已用范围延伸到这一行——这正是
    # 幽灵行产生的真实场景，不是凭空构造的极端情况。
    worksheet.append(["temp", "temp", "temp"])
    worksheet["A3"] = None
    worksheet["B3"] = None
    worksheet["C3"] = None
    workbook.save(str(tmp_path / "products.xlsx"))

    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.xlsx",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
        ],
        relations=[],
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.entities_written == 1
    assert report.entities_skipped == 0
    assert report.skipped_rows == []


async def test_standard_name_parts_are_joined_into_the_display_name(tmp_path):
    """两个同名不同邮编的客户，展示名带上判别列后可区分，且两条都落库。

    这正是 2026-08-30 那次事故要的结果：10000 行客户全部写入，而不是
    只写出 9335 条同名合并后的记录。
    """
    conn = await _confirmed_conn()
    (tmp_path / "customers.csv").write_text(
        "name,zip\nWilliam Jackson,72848\nWilliam Jackson,68046\n", encoding="utf-8"
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="customers.csv",
                standard_name_parts=["name", "zip"],
                node_key_parts=[ColumnNodeKeyPart(column="name"), ColumnNodeKeyPart(column="zip")],
                field_mappings={},
            ),
        ],
        relations=[],
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.entities_written == 2
    names = sorted(t.standard_name for t in await list_terms(conn, "muji"))
    assert names == ["William Jackson / 68046", "William Jackson / 72848"]


async def test_run_schema_etl_reads_utf8_bom_encoded_csv(tmp_path):
    """Excel 的"CSV UTF-8"导出格式会在文件开头写一个 BOM——见
    docs/superpowers/specs/2026-08-21-schema-etl-multi-format-upload.md
    决策 6 的补充说明。这里直接写带 BOM 的字节，验证读取器能正确剥掉
    BOM、不会让它粘在表头第一列名字前面导致列名对不上。"""
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_bytes(
        "﻿product_group_id,product_group_name,md_no\n1001,圆角收纳盒,A123\n".encode("utf-8")
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
        ],
        relations=[],
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.entities_written == 1
    assert report.entities_skipped == 0
    term = await get_term(conn, tenant_id="muji", standard_name="圆角收纳盒")
    assert term is not None
    assert term.node_key == "Product:1001"


async def test_run_schema_etl_raises_on_duplicate_node_keys_and_writes_nothing(tmp_path):
    """主键重复是配置错误，不是脏数据：node_key_parts 声明的列组合不足以
    唯一标识每一行。整体失败、零写入——部分写入会留下一个"看起来成功了、
    实际缺了一部分"的图谱，比失败更难发现。"""
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name,md_no\n"
        "P1,甲,M1\n"
        "P1,乙,M2\n",
        encoding="utf-8",
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
        ],
        relations=[],
    )
    graph_client = FakeGraphClient()

    with pytest.raises(DuplicateNodeKeyError) as excinfo:
        await run_schema_etl(
            conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
        )

    message = str(excinfo.value)
    assert "Product" in message
    assert "Product:P1" in message
    assert "2, 3" in message  # 冲突的源文件行号，第 1 行是表头
    assert await list_terms(conn, "muji") == []
    assert graph_client.synced == []


async def test_run_schema_etl_dirty_rows_still_skip_instead_of_failing_the_whole_run(tmp_path):
    """行级脏数据（缺列）语义不变：跳过 + 记报告，不升级成整体失败。
    只有主键重复才整体失败。"""
    conn = await _confirmed_conn()
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name,md_no\n"
        "P1,甲,M1\n"
        ",乙,M2\n",
        encoding="utf-8",
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={"md_no": "md_no"},
            ),
        ],
        relations=[],
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.entities_written == 1
    assert report.entities_skipped == 1
    assert len(await list_terms(conn, "muji")) == 1


async def test_run_schema_etl_reuses_stable_codes_allocated_before_a_duplicate_failure(tmp_path):
    """预检会调 compute_node_key(allow_allocation=True)，也就是说预检失败
    之前稳定码已经写进 etl_stable_code_registry 了——"零写入"的准确表述是
    "terms 和图零写入"，不是"零副作用"。这条用例钉住副作用是无害的：
    稳定码幂等分配，下次运行命中同一个码，不产生新码、不漂移。

    node_key_parts 只用 dup_key 这一列分配稳定码（两行的 dup_key 都是
    "K1"），variant_value（红/蓝）只用来生成 standard_name——这跟
    ColumnNodeKeyPart 版本里 product_group_id 重复、product_group_name
    不同的结构完全对应，只是键的来源换成了稳定码分配路径，测试才真的
    验证了自己名字承诺的"稳定码幂等"。"""
    conn = await _confirmed_conn()
    (tmp_path / "variants.csv").write_text(
        "variant_value,dup_key\n"
        "红,K1\n"
        "蓝,K1\n",
        encoding="utf-8",
    )
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="VariantValue", source_file="variants.csv",
                standard_name_parts=["variant_value"],
                node_key_parts=[
                    AllocatedCodeNodeKeyPart(scope_columns=[], raw_value_column="dup_key")
                ],
                field_mappings={},
            ),
        ],
        relations=[],
    )

    with pytest.raises(DuplicateNodeKeyError):
        await run_schema_etl(
            conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
        )
    code_after_first = await lookup_stable_code(
        conn, tenant_id="muji", scope="VariantValue", raw_value="K1"
    )
    assert code_after_first is not None

    with pytest.raises(DuplicateNodeKeyError):
        await run_schema_etl(
            conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
        )

    # 两次运行都失败，terms 始终是空的——重复的失败不会累积出半份数据。
    assert await list_terms(conn, "muji") == []
    # 稳定码幂等分配：第二次运行命中同一个码，不产生新码、不漂移。
    assert await lookup_stable_code(
        conn, tenant_id="muji", scope="VariantValue", raw_value="K1"
    ) == code_after_first


async def test_run_schema_etl_removes_entities_that_vanished_from_the_source(tmp_path):
    """源里删掉一行，重跑之后那个实体就该从 terms 和图谱里消失——数据源是
    权威的，本体是它的投影。ETL 此前只有 upsert、没有任何删除，源修正后
    得到的是新旧并存而不是修正后的状态。"""
    conn = await _confirmed_conn()
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={},
            ),
        ],
        relations=[],
    )
    path = tmp_path / "products.csv"
    path.write_text(
        "product_group_id,product_group_name\nP1,甲\nP2,乙\nP3,丙\n", encoding="utf-8"
    )
    await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )
    assert len(await list_terms(conn, "muji")) == 3

    path.write_text("product_group_id,product_group_name\nP1,甲\nP2,乙\n", encoding="utf-8")
    graph_client = FakeGraphClient()
    report = await run_schema_etl(
        conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
    )

    assert report.entities_removed == 1
    assert report.entities_removed_by_type == {"Product": 1}
    assert {t.node_key for t in await list_terms(conn, "muji")} == {"Product:P1", "Product:P2"}
    # 图谱侧也要删——delete_term_node 是 DETACH DELETE，连边和别名节点一起清。
    assert graph_client.deleted_nodes == ["Product:P3"]


async def test_run_schema_etl_sweep_never_touches_manually_created_terms(tmp_path):
    """审核界面创建的实体（source='review'）从来就不来自这个数据源，
    "源里没有"对它不成立。即使它的 term_type 由 ETL 管理，也不能被扫掉。"""
    conn = await _confirmed_conn()
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={},
            ),
        ],
        relations=[],
    )
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name\nP1,甲\n", encoding="utf-8"
    )
    await upsert_term_with_node_key(
        conn, tenant_id="muji", node_key="Product:HAND", standard_name="手工产品",
        aliases=[], term_type="Product", extra_properties={}, source="review",
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.entities_removed == 0
    assert "Product:HAND" in {t.node_key for t in await list_terms(conn, "muji")}


async def test_run_schema_etl_reports_zero_removals_explicitly(tmp_path):
    """零删除也要出现在报告里——"本次没有移除任何实体"和"根本没跑删除
    逻辑"必须能区分开。"""
    conn = await _confirmed_conn()
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={},
            ),
        ],
        relations=[],
    )
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name\nP1,甲\n", encoding="utf-8"
    )

    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    assert report.entities_removed == 0
    assert report.entities_removed_by_type == {"Product": 0}
    assert report.relations_removed == 0


async def test_run_schema_etl_safety_valve_aborts_with_zero_changes(tmp_path):
    """一次误传的、被截断的源文件会静默清空大半个图谱，而症状要等用户提问
    答不出来才暴露。阈值把最常见的事故形态挡在门外，且触发时整轮零改动——
    不做部分清理。"""
    conn = await _confirmed_conn()
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={},
            ),
        ],
        relations=[],
    )
    path = tmp_path / "products.csv"
    path.write_text(
        "product_group_id,product_group_name\nP1,甲\nP2,乙\nP3,丙\nP4,丁\n",
        encoding="utf-8",
    )
    await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )
    before = {t.node_key for t in await list_terms(conn, "muji")}
    assert len(before) == 4

    # 截断到只剩 1 行：将要移除 3/4 = 75%，超过 50% 阈值。
    path.write_text("product_group_id,product_group_name\nP1,甲\n", encoding="utf-8")
    graph_client = FakeGraphClient()

    with pytest.raises(SweepSafetyValveError) as excinfo:
        await run_schema_etl(
            conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
        )

    message = str(excinfo.value)
    assert "Product" in message
    assert "3" in message and "4" in message
    # 零改动：既没删，也没写。
    assert {t.node_key for t in await list_terms(conn, "muji")} == before
    assert graph_client.deleted_nodes == []
    assert graph_client.synced == []


async def test_run_schema_etl_allow_large_sweep_lets_the_run_through(tmp_path):
    """阈值是启发式，不是正确性保证。租户确实要缩减数据时必须有显式的放行
    方式，否则安全阀会把合法操作永久挡死。"""
    conn = await _confirmed_conn()
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={},
            ),
        ],
        relations=[],
    )
    path = tmp_path / "products.csv"
    path.write_text(
        "product_group_id,product_group_name\nP1,甲\nP2,乙\nP3,丙\nP4,丁\n",
        encoding="utf-8",
    )
    await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    path.write_text("product_group_id,product_group_name\nP1,甲\n", encoding="utf-8")
    report = await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config,
        data_dir=tmp_path, allow_large_sweep=True,
    )

    assert report.entities_removed == 3
    assert {t.node_key for t in await list_terms(conn, "muji")} == {"Product:P1"}


async def test_run_schema_etl_dry_run_reports_removals_without_changing_anything(tmp_path):
    """首次启用 sweep 会清理掉历史累积的孤儿实体，规模可能不小。dry-run 让
    租户先看一眼将要删什么，再决定是否真跑。"""
    conn = await _confirmed_conn()
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={},
            ),
        ],
        relations=[],
    )
    path = tmp_path / "products.csv"
    path.write_text(
        "product_group_id,product_group_name\nP1,甲\nP2,乙\n", encoding="utf-8"
    )
    await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    path.write_text("product_group_id,product_group_name\nP1,甲\n", encoding="utf-8")
    graph_client = FakeGraphClient()
    report = await run_schema_etl(
        conn=conn, graph_client=graph_client, config=config,
        data_dir=tmp_path, dry_run=True,
    )

    assert report.dry_run is True
    assert report.entities_removed == 1
    assert report.entities_removed_by_type == {"Product": 1}
    # 什么都没动。
    assert len(await list_terms(conn, "muji")) == 2
    assert graph_client.deleted_nodes == []
    assert graph_client.synced == []


def _product_sku_config() -> SchemaETLConfig:
    """Product/SKU 各一条实体映射、一条 HAS_SKU 关系映射，三者都来自
    同一个 products.csv——用于验证关系"先写后扫"只需要一个源文件。"""
    return SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={},
            ),
            EntityMapping(
                term_type="SKU", source_file="products.csv",
                standard_name_parts=["jan"],
                node_key_parts=[ColumnNodeKeyPart(column="jan")],
                field_mappings={},
            ),
        ],
        relations=[
            RelationMapping(
                relation_type="HAS_SKU", source_file="products.csv",
                subject_term_type="Product", object_term_type="SKU",
            ),
        ],
    )


def _two_relations_same_source_config() -> SchemaETLConfig:
    """Product/SKU/VariantValue 三条实体映射 + HAS_SKU/HAS_VARIANT_VALUE 两条
    关系映射，全部指向同一个 products.csv——对应简报里"demo 配置里五条关系
    全部来自同一个源文件"的场景，用来验证按源文件去重而不是按映射逐一扫。
    HAS_VARIANT_VALUE 的主客体都是 VariantValue，两端键算出来天然相同，
    这里只是为了触发写入+扫除，不追求业务上合理。"""
    return SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={},
            ),
            EntityMapping(
                term_type="SKU", source_file="products.csv",
                standard_name_parts=["jan"],
                node_key_parts=[ColumnNodeKeyPart(column="jan")],
                field_mappings={},
            ),
            EntityMapping(
                term_type="VariantValue", source_file="products.csv",
                standard_name_parts=["variant_value"],
                node_key_parts=[ColumnNodeKeyPart(column="variant_value")],
                field_mappings={},
            ),
        ],
        relations=[
            RelationMapping(
                relation_type="HAS_SKU", source_file="products.csv",
                subject_term_type="Product", object_term_type="SKU",
            ),
            RelationMapping(
                relation_type="HAS_VARIANT_VALUE", source_file="products.csv",
                subject_term_type="VariantValue", object_term_type="VariantValue",
            ),
        ],
    )


def _write_product_sku_source(tmp_path: Path) -> None:
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name,jan,variant_value\n"
        "1001,圆角收纳盒,4901234567890,红色\n",
        encoding="utf-8",
    )


async def test_run_schema_etl_sweeps_stale_relations_after_writing_fresh_ones(tmp_path):
    """关系用"先写新边、再扫陈旧边"：任何时刻图谱都是完整的，中途失败最多
    留下新旧共存，下次重跑自愈。若改成"先全删再全写"，中途失败会留下一个
    边被删光、实体还在的图谱，而 ETL 数据量大、这个窗口很长。"""
    conn = await _confirmed_conn()
    config = _product_sku_config()
    _write_product_sku_source(tmp_path)
    graph_client = FakeGraphClient()

    await run_schema_etl(
        conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
    )

    # 扫除发生了，且针对配置里出现过的源文件。
    assert [s for s, _ in graph_client.stale_sweeps] == ["products.csv"]
    # 扫除的时间界线就是本轮的写入时间——本轮写的边一律不会被扫掉。
    assert graph_client.stale_sweeps[0][1]


async def test_run_schema_etl_sweeps_each_source_file_once_not_once_per_mapping(tmp_path):
    """多条关系映射共享同一个源文件时（demo 配置里五条关系全部来自
    soft_drink_sales.xlsx），扫除必须按源文件去重，不能每个映射扫一遍。"""
    conn = await _confirmed_conn()
    config = _two_relations_same_source_config()
    _write_product_sku_source(tmp_path)
    graph_client = FakeGraphClient()

    await run_schema_etl(
        conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
    )

    assert [s for s, _ in graph_client.stale_sweeps] == ["products.csv"]
    # 两种关系都写进去了，没有互相抹掉。
    assert {r for _, _, r in graph_client.merged} == {"HAS_SKU", "HAS_VARIANT_VALUE"}


async def test_run_schema_etl_rejects_duplicate_entity_term_type(tmp_path):
    """config.entities 里两条映射共享同一个 term_type 是配置错误，不是可以
    逐行跳过的脏数据：run_schema_etl 内部两处（entity_mappings_by_term_type、
    scanned_keys_by_term_type）都按 term_type 建字典，重复声明会被静默折叠
    成后一条，前一条映射写入的实体会在下一轮被误判成"源里已经消失"进而被
    sweep 删除——代价从"关系端点查找不准"升级成了静默数据丢失。必须在任何
    写入之前整体拒绝、零改动。"""
    conn = await _confirmed_conn()
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products_a.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={},
            ),
            EntityMapping(
                term_type="Product", source_file="products_b.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={},
            ),
        ],
        relations=[],
    )
    graph_client = FakeGraphClient()

    with pytest.raises(DuplicateEntityMappingError):
        await run_schema_etl(
            conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
        )

    assert len(await list_terms(conn, "muji")) == 0
    assert graph_client.synced == []
    assert graph_client.deleted_nodes == []


async def test_run_schema_etl_sweep_deletes_graph_node_before_sqlite_row(tmp_path):
    """双存储删除顺序决定了中途失败能不能自愈：必须先删 Neo4j 节点、再删
    SQLite 行。doomed 集合本身算自 SQLite 的 source='etl' 行，如果反过来
    先删 SQLite（内部已 commit）、再删 Neo4j，delete_term_node 中途抛异常
    时 SQLite 行已经没了，留下一个再也不会进入任何一次未来 sweep 候选集
    的孤儿节点——不可自愈。现在这个顺序下，同样的异常发生时 SQLite 还
    完好，下次重跑会重新算出同一个 doomed 集合、对着同一个节点再调一次
    delete_term_node（MATCH 不到是空操作，天然幂等），最终收敛。"""
    conn = await _confirmed_conn()
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={},
            ),
        ],
        relations=[],
    )
    path = tmp_path / "products.csv"
    path.write_text(
        "product_group_id,product_group_name\nP1,甲\nP2,乙\n", encoding="utf-8"
    )
    await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    # P2 从源里消失，下一轮会被 sweep 判定为 doomed。
    path.write_text("product_group_id,product_group_name\nP1,甲\n", encoding="utf-8")

    class FailingGraphClient(FakeGraphClient):
        async def delete_term_node(self, *, tenant_id: str, node_key: str) -> None:
            raise RuntimeError("图数据库暂时不可用")

    graph_client = FailingGraphClient()

    with pytest.raises(RuntimeError):
        await run_schema_etl(
            conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
        )

    # 图删在前、SQLite 删在后：异常发生在图删除这一步，此刻 SQLite 侧的
    # 批量删除还没执行——Product:P2 这一行必须还在，这就是可自愈方向的
    # 证据：只要 SQLite 行还在，下次重跑就还会把它算进同一个 doomed 集合。
    assert "Product:P2" in {t.node_key for t in await list_terms(conn, "muji")}


async def test_run_schema_etl_skips_relation_row_whose_endpoint_is_doomed_this_round(tmp_path):
    """关系写入先于实体 sweep：如果不扣掉本轮已判定为 doomed 的端点，一条
    关系行可能引用一个"上一轮写的、本轮源里已经消失"的实体——写入时刻它
    在 SQLite 里还没被删（sweep 在写入之后才执行），端点存在性守卫会误判
    成"端点还在"而放行，紧接着 sweep 就把这个刚写的端点连同这条边一起
    DETACH DELETE 掉。relations_written 计过数、边却不存在，这是写完立刻
    删的不自洽。扣掉 doomed 之后，这一行应该走既有的"端点不存在"路径，
    被跳过并记进 skipped_rows。"""
    conn = await _confirmed_conn()
    config = SchemaETLConfig(
        tenant_id="muji",
        entities=[
            EntityMapping(
                term_type="Product", source_file="products.csv",
                standard_name_parts=["product_group_name"],
                node_key_parts=[ColumnNodeKeyPart(column="product_group_id")],
                field_mappings={},
            ),
            EntityMapping(
                term_type="SKU", source_file="skus.csv",
                standard_name_parts=["jan"],
                node_key_parts=[ColumnNodeKeyPart(column="jan")],
                field_mappings={},
            ),
        ],
        relations=[
            RelationMapping(
                relation_type="HAS_SKU", source_file="skus.csv",
                subject_term_type="Product", object_term_type="SKU",
            ),
        ],
    )
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name\nP1,甲\nP2,乙\n", encoding="utf-8"
    )
    (tmp_path / "skus.csv").write_text(
        "jan,product_group_id\nS1,P1\n", encoding="utf-8"
    )
    await run_schema_etl(
        conn=conn, graph_client=FakeGraphClient(), config=config, data_dir=tmp_path
    )

    # Product:P1 从实体源里消失（本轮会被判定为 doomed），但 skus.csv 没
    # 变——它引用的 product_group_id 仍然是 P1，关系行还会尝试写这条边。
    (tmp_path / "products.csv").write_text(
        "product_group_id,product_group_name\nP2,乙\n", encoding="utf-8"
    )
    graph_client = FakeGraphClient()
    report = await run_schema_etl(
        conn=conn, graph_client=graph_client, config=config, data_dir=tmp_path
    )

    assert report.entities_removed_by_type["Product"] == 1
    # 这一行被跳过，不再计入 relations_written，也没有被写进图谱。
    assert report.relations_written == 0
    assert report.relations_skipped == 1
    assert ("Product:P1", "SKU:S1", "HAS_SKU") not in graph_client.merged
    skipped = [r for r in report.skipped_rows if r.label == "HAS_SKU"]
    assert len(skipped) == 1
    assert "Product:P1" in skipped[0].reason
