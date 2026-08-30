"""ETL 三层管道的第一层：staging——解析与类型归一。

xlsx/csv/tsv/xls → 统一的 dict[str, str] 行序列。这一层不知道本体、
不知道 node_key，只负责"把文件变成行"。2026-08-30 从 schema_etl.py
原样提取，行为一行未改。
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterator

import xlrd
from openpyxl import load_workbook

from app.graphrag.schema_etl_row_processing import RowProcessingError, convert_excel_cell_to_string


def _detect_text_encoding(path: Path) -> str:
    """CSV/TSV 源文件的编码探测：优先按 UTF-8 严格解码，失败则回退尝试
    GBK（国内 Excel 导出 CSV 最常见的默认编码）——见
    docs/superpowers/specs/2026-08-21-schema-etl-multi-format-upload.md
    决策 6。这里读一遍原始字节只是为了做 decode 测试，不保留解码结果；
    真正的行级处理仍然通过 csv.DictReader 用确定的编码重新打开文件、
    流式进行，不会把整份解码后的文本一次性留在内存里。两种编码都解码
    失败时，让 GBK 阶段的 UnicodeDecodeError 原样往上抛，不做进一步猜测。

    UTF-8 分支返回 "utf-8-sig" 而不是 "utf-8"：Excel 的"CSV UTF-8"导出
    格式会在文件开头写一个 BOM，纯 "utf-8" 编码不会因为这个 BOM 报解码
    错误，但会把它解码成字面的 U+FEFF 字符粘在第一个字段名前面，导致
    表头第一列对不上用户看到的列名、整份文件被误判成"缺列"全部跳过。
    "utf-8-sig" 对没有 BOM 的普通 UTF-8 文件解码结果完全一样，只在文件
    真的带 BOM 时才会正确剥掉它，是纯粹的超集写法。
    """
    raw = path.read_bytes()
    try:
        raw.decode("utf-8")
        return "utf-8-sig"
    except UnicodeDecodeError:
        pass
    raw.decode("gbk")
    return "gbk"


def _read_delimited_rows(path: Path, *, delimiter: str) -> Iterator[dict[str, str]]:
    """逐行流式产出 CSV/TSV 源文件的行——设计文档第 6.4 节给出的真实规模
    是"MUJI 一张 SKU 表 18 万+ 行"。注意：编码探测阶段（见
    _detect_text_encoding）会把整个文件读一遍原始字节做 decode 测试，
    这一步不是流式的；探测完成后的逐行处理本身才是流式、不整份文本
    常驻内存。"""
    encoding = _detect_text_encoding(path)
    with path.open(encoding=encoding, newline="") as handle:
        yield from csv.DictReader(handle, delimiter=delimiter)


def _read_xlsx_rows(path: Path) -> Iterator[dict[str, str]]:
    """流式读取 xlsx 第一个工作表，第一行是表头——见决策 2（固定读第一个
    sheet）。read_only=True 让 openpyxl 用懒加载模式逐行产出，不把整个
    工作表读进内存，跟 CSV 路径同一个"18 万+ 行不能爆内存"的约束。
    data_only=True 拿单元格公式算出来的值，不拿公式字符串本身。"""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return
        header = [str(cell).strip() if cell is not None else "" for cell in header_row]
        for row in rows_iter:
            values = {
                header[i]: convert_excel_cell_to_string(row[i] if i < len(row) else None)
                for i in range(len(header))
            }
            # openpyxl 的 read_only 迭代会按工作表"已用范围"补齐行数，哪怕
            # 某一行早就被清空也会产出全空的幽灵行（常见于手工编辑过的
            # Excel 导出文件）。跳过全空行，避免这些幽灵行被当成"缺列"的
            # 脏数据行计入 skipped_rows，也让这条路径跟 CSV 侧
            # csv.DictReader 对空行的处理保持一致。
            if not any(values.values()):
                continue
            yield values
    finally:
        workbook.close()


def _xlrd_cell_to_python_value(cell: "xlrd.sheet.Cell", datemode: int) -> object:
    """把 xlrd 的 Cell（用 ctype 标记类型、日期存成 Excel 序列号）归一化成
    openpyxl 风格的原生 Python 值（int/float/str/bool/datetime/None），
    这样就能复用同一个 convert_excel_cell_to_string 做字符串化，不用给
    xlrd 单独写一套转换规则。"""
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, datemode)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        # 公式错误单元格（#DIV/0!、#N/A 等）——cell.value 是一个内部错误码
        # 整数，直接透传会被字符串化成一个看起来正常、实际语义错误的值。
        # 当成空值处理，让这一行走"缺列"的正常脏数据路径。
        return None
    return cell.value  # XL_CELL_NUMBER（float）/ XL_CELL_TEXT（str）


def _read_xls_rows(path: Path) -> Iterator[dict[str, str]]:
    """读取旧版二进制 xls 第一个工作表，第一行是表头。xlrd 没有 openpyxl
    那种懒加载流式模式，会把整个工作表读进内存——xls 是被淘汰的旧格式，
    体量通常不大，这里不为了流式特意做额外处理。"""
    workbook = xlrd.open_workbook(str(path))
    worksheet = workbook.sheet_by_index(0)
    if worksheet.nrows == 0:
        return
    header = [str(worksheet.cell_value(0, col)).strip() for col in range(worksheet.ncols)]
    for row_idx in range(1, worksheet.nrows):
        values = {
            header[col_idx]: convert_excel_cell_to_string(
                _xlrd_cell_to_python_value(worksheet.cell(row_idx, col_idx), workbook.datemode)
            )
            for col_idx in range(len(header))
        }
        # 跟 _read_xlsx_rows 同样的理由：跳过全空行。
        if not any(values.values()):
            continue
        yield values


def read_table_rows(path: Path) -> Iterator[dict[str, str]]:
    """按扩展名分流到对应的行读取器，统一产出 dict[str, str]——见
    docs/superpowers/specs/2026-08-21-schema-etl-multi-format-upload.md。

    这是 ETL 三层管道的第一层（staging）的唯一入口：解析 + 类型归一，
    不认识 node_key、不认识本体，只把各种格式的表统一成行序列。见
    docs/superpowers/specs/2026-08-30-etl-layered-pipeline-design.md。
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from _read_delimited_rows(path, delimiter=",")
    elif suffix == ".tsv":
        yield from _read_delimited_rows(path, delimiter="\t")
    elif suffix == ".xlsx":
        yield from _read_xlsx_rows(path)
    elif suffix == ".xls":
        yield from _read_xls_rows(path)
    else:
        raise RowProcessingError(f"不支持的数据文件类型: {suffix!r}（{path.name}）")
