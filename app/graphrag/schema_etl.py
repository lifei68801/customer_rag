from __future__ import annotations

import argparse
import asyncio
import csv
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterator, Protocol

import aiosqlite
import xlrd
from openpyxl import load_workbook

from app.config.settings import Settings
from app.graphrag import provenance
from app.graphrag.etl_stable_code_registry import ensure_stable_code_registry_schema
from app.graphrag.factory import build_graph_client_from_settings
from app.graphrag.ontology import Term
from app.graphrag.relation_writer import RelationWriterProtocol
from app.graphrag.ontology_categories import list_term_types
from app.graphrag.ontology_constraints import list_allowed_combinations, to_combination_keys
from app.graphrag.ontology_lifecycle import is_ontology_confirmed
from app.graphrag.ontology_relations import list_relation_types
from app.graphrag.ontology_store import open_ontology_store_conn
from app.graphrag.schema_etl_config import EntityMapping, RelationMapping, SchemaETLConfig, load_schema_etl_config
from app.graphrag.schema_etl_row_processing import (
    RowProcessingError,
    compute_node_key,
    convert_excel_cell_to_string,
    convert_field_value,
)
from app.graphrag.terms_store import (
    TermNameConflictError,
    UnknownCategoryError,
    list_node_keys_by_term_type,
    upsert_term_with_node_key,
)


class SchemaEtlGraphProtocol(RelationWriterProtocol, Protocol):
    """ETL 引擎实际调用的两个图写方法——独立声明，不继承 admin 路由用的
    GraphWriteProtocol（neo4j_client.py）或摄取管道用的
    GraphWriteClientProtocol（normalization.py）：三者是不同消费方，各自
    只暴露自己真正用到的方法，见 2026-08-27 架构评审对 GraphClientProtocol
    过宽问题的讨论。merge_relation 的签名继承自 RelationWriterProtocol——
    窄协议的意图不变，只是那份签名不再各抄一遍。"""

    async def sync_term(self, term: Term) -> None: ...


class SchemaETLNotConfirmedError(Exception):
    """该租户的本体 schema 还没有 confirm，拒绝运行 ETL——见
    docs/superpowers/specs/2026-08-16-schema-etl-engine-design.md 第 6.2 节。"""


@dataclass
class SkippedRow:
    label: str
    source_file: str
    row_number: int
    reason: str


@dataclass
class SkippedMapping:
    label: str
    source_file: str
    reason: str


@dataclass
class ETLRunReport:
    entities_written: int = 0
    entities_skipped: int = 0
    relations_written: int = 0
    relations_skipped: int = 0
    written_by_type: dict[str, int] = field(default_factory=dict)
    skipped_by_type: dict[str, int] = field(default_factory=dict)
    skipped_rows: list[SkippedRow] = field(default_factory=list)
    skipped_mappings: list[SkippedMapping] = field(default_factory=list)


def _detect_text_encoding(path: Path) -> str:
    """CSV/TSV 源文件的编码探测：优先按 UTF-8 严格解码，失败则回退尝试
    GBK（国内 Excel 导出 CSV 最常见的默认编码）——见
    docs/superpowers/specs/2026-08-21-schema-etl-multi-format-upload.md
    决策 6。这里读一遍原始字节只是为了做 decode 测试，不保留解码结果；
    真正的行级处理仍然通过 csv.DictReader 用确定的编码重新打开文件、
    流式进行，不会把整份解码后的文本一次性留在内存里。两种编码都解码
    失败时，让 GBK 阶段的 UnicodeDecodeError 原样往上抛，不做进一步猜测。

    UTF-8 分支返回 "utf-8-sig" 而不是 "utf-8"：Excel 的"CSV UTF-8"导出
    格式会在文件开头写一个 BOM，纯 "utf-8" 编码不会因为这个 BOM 报解码
    错误，但会把它解码成字面的 U+FEFF 字符粘在第一个字段名前面，导致
    表头第一列对不上用户看到的列名、整份文件被误判成"缺列"全部跳过。
    "utf-8-sig" 对没有 BOM 的普通 UTF-8 文件解码结果完全一样，只在文件
    真的带 BOM 时才会正确剥掉它，是纯粹的超集写法。
    """
    raw = path.read_bytes()
    try:
        raw.decode("utf-8")
        return "utf-8-sig"
    except UnicodeDecodeError:
        pass
    raw.decode("gbk")
    return "gbk"


def _read_delimited_rows(path: Path, *, delimiter: str) -> Iterator[dict[str, str]]:
    """逐行流式产出 CSV/TSV 源文件的行——设计文档第 6.4 节给出的真实规模
    是"MUJI 一张 SKU 表 18 万+ 行"。注意：编码探测阶段（见
    _detect_text_encoding）会把整个文件读一遍原始字节做 decode 测试，
    这一步不是流式的；探测完成后的逐行处理本身才是流式、不整份文本
    常驻内存。"""
    encoding = _detect_text_encoding(path)
    with path.open(encoding=encoding, newline="") as handle:
        yield from csv.DictReader(handle, delimiter=delimiter)


def _read_xlsx_rows(path: Path) -> Iterator[dict[str, str]]:
    """流式读取 xlsx 第一个工作表，第一行是表头——见决策 2（固定读第一个
    sheet）。read_only=True 让 openpyxl 用懒加载模式逐行产出，不把整个
    工作表读进内存，跟 CSV 路径同一个"18 万+ 行不能爆内存"的约束。
    data_only=True 拿单元格公式算出来的值，不拿公式字符串本身。"""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return
        header = [str(cell).strip() if cell is not None else "" for cell in header_row]
        for row in rows_iter:
            values = {
                header[i]: convert_excel_cell_to_string(row[i] if i < len(row) else None)
                for i in range(len(header))
            }
            # openpyxl 的 read_only 迭代会按工作表"已用范围"补齐行数，哪怕
            # 某一行早就被清空也会产出全空的幽灵行（常见于手工编辑过的
            # Excel 导出文件）。跳过全空行，避免这些幽灵行被当成"缺列"的
            # 脏数据行计入 skipped_rows，也让这条路径跟 CSV 侧
            # csv.DictReader 对空行的处理保持一致。
            if not any(values.values()):
                continue
            yield values
    finally:
        workbook.close()


def _xlrd_cell_to_python_value(cell: "xlrd.sheet.Cell", datemode: int) -> object:
    """把 xlrd 的 Cell（用 ctype 标记类型、日期存成 Excel 序列号）归一化成
    openpyxl 风格的原生 Python 值（int/float/str/bool/datetime/None），
    这样就能复用同一个 convert_excel_cell_to_string 做字符串化，不用给
    xlrd 单独写一套转换规则。"""
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return None
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate_as_datetime(cell.value, datemode)
    if cell.ctype == xlrd.XL_CELL_ERROR:
        # 公式错误单元格（#DIV/0!、#N/A 等）——cell.value 是一个内部错误码
        # 整数，直接透传会被字符串化成一个看起来正常、实际语义错误的值。
        # 当成空值处理，让这一行走"缺列"的正常脏数据路径。
        return None
    return cell.value  # XL_CELL_NUMBER（float）/ XL_CELL_TEXT（str）


def _read_xls_rows(path: Path) -> Iterator[dict[str, str]]:
    """读取旧版二进制 xls 第一个工作表，第一行是表头。xlrd 没有 openpyxl
    那种懒加载流式模式，会把整个工作表读进内存——xls 是被淘汰的旧格式，
    体量通常不大，这里不为了流式特意做额外处理。"""
    workbook = xlrd.open_workbook(str(path))
    worksheet = workbook.sheet_by_index(0)
    if worksheet.nrows == 0:
        return
    header = [str(worksheet.cell_value(0, col)).strip() for col in range(worksheet.ncols)]
    for row_idx in range(1, worksheet.nrows):
        values = {
            header[col_idx]: convert_excel_cell_to_string(
                _xlrd_cell_to_python_value(worksheet.cell(row_idx, col_idx), workbook.datemode)
            )
            for col_idx in range(len(header))
        }
        # 跟 _read_xlsx_rows 同样的理由：跳过全空行。
        if not any(values.values()):
            continue
        yield values


def _read_table_rows(path: Path) -> Iterator[dict[str, str]]:
    """按扩展名分流到对应的行读取器，统一产出 dict[str, str]——见
    docs/superpowers/specs/2026-08-21-schema-etl-multi-format-upload.md。"""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        yield from _read_delimited_rows(path, delimiter=",")
    elif suffix == ".tsv":
        yield from _read_delimited_rows(path, delimiter="\t")
    elif suffix == ".xlsx":
        yield from _read_xlsx_rows(path)
    elif suffix == ".xls":
        yield from _read_xls_rows(path)
    else:
        raise RowProcessingError(f"不支持的数据文件类型: {suffix!r}（{path.name}）")


def _record_written(report: ETLRunReport, *, label: str) -> None:
    report.written_by_type[label] = report.written_by_type.get(label, 0) + 1


def _record_skipped_row(
    report: ETLRunReport, *, label: str, source_file: str, row_number: int, reason: str
) -> None:
    report.skipped_by_type[label] = report.skipped_by_type.get(label, 0) + 1
    report.skipped_rows.append(
        SkippedRow(label=label, source_file=source_file, row_number=row_number, reason=reason)
    )


async def _write_entity_mapping(
    *,
    conn: aiosqlite.Connection,
    graph_client: SchemaEtlGraphProtocol,
    tenant_id: str,
    mapping: EntityMapping,
    data_dir: Path,
    report: ETLRunReport,
) -> None:
    term_types = await list_term_types(conn, tenant_id, status="confirmed")
    types_by_value = {t.value: t for t in term_types}
    if mapping.term_type not in types_by_value:
        raise RowProcessingError(f"term_type {mapping.term_type!r} 不在已确认 schema 里")
    extra_field_specs = {f.name: f for f in types_by_value[mapping.term_type].extra_fields}

    for row_number, row in enumerate(_read_table_rows(data_dir / mapping.source_file), start=2):  # 第 1 行是表头
        try:
            node_key = await compute_node_key(
                conn, tenant_id=tenant_id, term_type=mapping.term_type,
                node_key_parts=mapping.node_key_parts, row=row,
            )
            missing = [c for c in mapping.standard_name_parts if not row.get(c)]
            if missing:
                raise RowProcessingError(
                    f"standard_name 需要的列 {missing!r} 不存在或为空"
                )
            # 用 " / " 连接，不用冒号——冒号是 node_key 的分隔符，展示名里
            # 再用一次会让两者在日志和界面上难以区分。
            standard_name = " / ".join(row[c] for c in mapping.standard_name_parts)
            extra_properties = {
                field_name: convert_field_value(
                    extra_field_specs=extra_field_specs, field_name=field_name,
                    raw_value=row[source_column],
                )
                for field_name, source_column in mapping.field_mappings.items()
                if source_column in row and row[source_column]
            }
            await upsert_term_with_node_key(
                conn, tenant_id=tenant_id, node_key=node_key, standard_name=standard_name,
                aliases=[], term_type=mapping.term_type,
                extra_properties=extra_properties,
            )
            term = Term(
                tenant_id=tenant_id, node_key=node_key, standard_name=standard_name,
                aliases=[], term_type=mapping.term_type,
                extra_properties=extra_properties,
            )
            await graph_client.sync_term(term)
            report.entities_written += 1
            _record_written(report, label=mapping.term_type)
        except (RowProcessingError, TermNameConflictError, UnknownCategoryError) as exc:
            report.entities_skipped += 1
            _record_skipped_row(
                report, label=mapping.term_type, source_file=mapping.source_file,
                row_number=row_number, reason=str(exc),
            )


async def _write_relation_mapping(
    *,
    conn: aiosqlite.Connection,
    graph_client: SchemaEtlGraphProtocol,
    tenant_id: str,
    mapping: RelationMapping,
    entity_mappings_by_term_type: dict[str, EntityMapping],
    confirmed_relation_types: set[str],
    allowed_combinations: set[tuple[str, str, str]],
    recorded_at: datetime,
    data_dir: Path,
    report: ETLRunReport,
) -> None:
    if mapping.relation_type not in confirmed_relation_types:
        raise RowProcessingError(f"relation_type {mapping.relation_type!r} 不在已确认 schema 里")
    # relation_type 单独合法不代表 (subject_term_type, relation_type,
    # object_term_type) 这个组合也在已确认的允许列表里——两者是独立声明的
    # 字段，映射配置本身不保证组合有效。这条校验让 ETL 写入路径追平
    # graph_extraction.py/review_queue.py 已经在做的同一种组合校验，见
    # docs/superpowers/specs/2026-08-19-data-entry-unification-design.md
    # "不在本次范围内"第 1 条的后续处理。
    combo = (mapping.subject_term_type, mapping.relation_type, mapping.object_term_type)
    if combo not in allowed_combinations:
        raise RowProcessingError(
            f"关系类型/实体类型组合不在已确认允许列表里: "
            f"({mapping.subject_term_type!r}, {mapping.relation_type!r}, {mapping.object_term_type!r})"
        )
    subject_entity = entity_mappings_by_term_type.get(mapping.subject_term_type)
    object_entity = entity_mappings_by_term_type.get(mapping.object_term_type)
    if subject_entity is None or object_entity is None:
        raise RowProcessingError(
            f"关系 {mapping.relation_type!r} 引用的实体类型未在 entities 段声明"
        )

    # 端点实体必须真的写进术语表过，否则跳过这一行——merge_relation 的两端
    # 都是 MERGE，node_key 对不上任何已有节点时不会报错，而是凭空建出一个
    # 只有 tenant_id/node_key、没有 type/standard_name 的幽灵节点（见
    # neo4j_client.py::merge_relation 的说明）。AllocatedCodeNodeKeyPart
    # 早就有这个守卫（稳定码查不到就跳过），ColumnNodeKeyPart 一直没有：
    # 实体行被跳过（唯一索引冲突、类型转换失败、缺列）时，关系行照写不误。
    # demo 租户就这样留下了 `类目:Coffee` 和 `销量:1000` 两个幽灵节点、
    # 挂着 16 条边。run_schema_etl 保证全部实体映射先于关系映射执行，所以
    # 这里预取一次就够，不会漏掉后面才写入的实体。
    subject_node_keys = await list_node_keys_by_term_type(
        conn, tenant_id, mapping.subject_term_type
    )
    object_node_keys = await list_node_keys_by_term_type(
        conn, tenant_id, mapping.object_term_type
    )

    for row_number, row in enumerate(_read_table_rows(data_dir / mapping.source_file), start=2):
        try:
            subject_key = await compute_node_key(
                conn, tenant_id=tenant_id, term_type=mapping.subject_term_type,
                node_key_parts=subject_entity.node_key_parts, row=row, allow_allocation=False,
            )
            object_key = await compute_node_key(
                conn, tenant_id=tenant_id, term_type=mapping.object_term_type,
                node_key_parts=object_entity.node_key_parts, row=row, allow_allocation=False,
            )
            for key, known_keys, term_type in (
                (subject_key, subject_node_keys, mapping.subject_term_type),
                (object_key, object_node_keys, mapping.object_term_type),
            ):
                if key not in known_keys:
                    raise RowProcessingError(
                        f"关系端点 {key!r} 在术语表里不存在"
                        f"（{term_type!r} 的实体行可能被跳过或尚未写入）"
                    )
            await graph_client.merge_relation(
                subject_standard_name=subject_key, object_standard_name=object_key,
                relation_type=mapping.relation_type, source=mapping.source_file,
                tenant_id=tenant_id, provenance=provenance.ETL, recorded_at=recorded_at,
            )
            report.relations_written += 1
            _record_written(report, label=mapping.relation_type)
        except RowProcessingError as exc:
            report.relations_skipped += 1
            _record_skipped_row(
                report, label=mapping.relation_type, source_file=mapping.source_file,
                row_number=row_number, reason=str(exc),
            )


async def run_schema_etl(
    *, conn: aiosqlite.Connection, graph_client: SchemaEtlGraphProtocol, config: SchemaETLConfig, data_dir: Path
) -> ETLRunReport:
    """按已确认 schema + 列映射配置，把 CSV 源数据确定性写入 Term/Neo4j 双存储。
    见 docs/superpowers/specs/2026-08-16-schema-etl-engine-design.md 第 6 节。

    单个 mapping 级别的前置校验失败（term_type/relation_type 不在已确认 schema
    里、关系引用了未声明的实体类型）不会中断整次运行——跳过这一个 mapping、
    记入 skipped_mappings，继续处理其余 mapping，呼应第 6.4 节"一行脏数据不该
    让整批任务失败"的同一原则，只是粒度提升到了整个 mapping。
    """
    if not await is_ontology_confirmed(conn, config.tenant_id):
        raise SchemaETLNotConfirmedError(
            f"租户 {config.tenant_id!r} 的本体 schema 还没有确认，拒绝运行 ETL"
        )
    await ensure_stable_code_registry_schema(conn)

    recorded_at = datetime.now()
    report = ETLRunReport()

    for entity_mapping in config.entities:
        try:
            await _write_entity_mapping(
                conn=conn, graph_client=graph_client, tenant_id=config.tenant_id,
                mapping=entity_mapping, data_dir=data_dir, report=report,
            )
        except RowProcessingError as exc:
            report.skipped_mappings.append(
                SkippedMapping(
                    label=entity_mapping.term_type, source_file=entity_mapping.source_file, reason=str(exc),
                )
            )

    confirmed_relation_types = {
        r.relation_type for r in await list_relation_types(conn, config.tenant_id, status="confirmed")
    }
    allowed_combinations = to_combination_keys(
        await list_allowed_combinations(conn, config.tenant_id, status="confirmed")
    )
    entity_mappings_by_term_type = {m.term_type: m for m in config.entities}
    for relation_mapping in config.relations:
        try:
            await _write_relation_mapping(
                conn=conn, graph_client=graph_client, tenant_id=config.tenant_id,
                mapping=relation_mapping, entity_mappings_by_term_type=entity_mappings_by_term_type,
                confirmed_relation_types=confirmed_relation_types,
                allowed_combinations=allowed_combinations, recorded_at=recorded_at,
                data_dir=data_dir, report=report,
            )
        except RowProcessingError as exc:
            report.skipped_mappings.append(
                SkippedMapping(
                    label=relation_mapping.relation_type, source_file=relation_mapping.source_file, reason=str(exc),
                )
            )

    return report


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="按列映射配置把结构化 CSV 数据写入知识图谱")
    parser.add_argument("--config", required=True, type=Path, help="列映射 YAML 配置文件路径")
    parser.add_argument("--data-dir", required=True, type=Path, help="配置里 source_file 相对路径的基准目录")
    return parser.parse_args()


async def _main(*, config_path: Path, data_dir: Path) -> None:
    settings = Settings()
    config = load_schema_etl_config(config_path)
    conn = await open_ontology_store_conn(settings)
    graph_client = build_graph_client_from_settings(settings)
    try:
        report = await run_schema_etl(conn=conn, graph_client=graph_client, config=config, data_dir=data_dir)
    finally:
        await conn.close()
    print(
        f"实体写入 {report.entities_written} 条，跳过 {report.entities_skipped} 条；"
        f"关系写入 {report.relations_written} 条，跳过 {report.relations_skipped} 条"
    )
    for skipped in report.skipped_mappings:
        print(f"  跳过整个映射 {skipped.label}（{skipped.source_file}）：{skipped.reason}")
    for skipped in report.skipped_rows:
        print(f"  跳过 {skipped.label} / {skipped.source_file} 第 {skipped.row_number} 行：{skipped.reason}")


if __name__ == "__main__":
    args = _parse_args()
    asyncio.run(_main(config_path=args.config, data_dir=args.data_dir))
